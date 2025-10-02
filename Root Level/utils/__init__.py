```python
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import io
import hashlib
import re
from typing import Dict, Any, Tuple

class SecureFileHandler:
    """Secure file upload and processing handler with comprehensive validation"""
    
    def __init__(self):
        self.MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
        self.ALLOWED_EXTENSIONS = ['.xlsx']
        self.ALLOWED_MIME_TYPES = [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ]
    
    def validate_and_process(self, uploaded_file) -> Dict[str, Any]:
        """
        Comprehensive validation and processing of uploaded Excel files
        Returns: Dictionary with success status, data, and messages
        """
        try:
            # Step 1: Basic file validation
            validation_result = self._validate_file(uploaded_file)
            if not validation_result['valid']:
                return {
                    'success': False,
                    'message': validation_result['message'],
                    'data': None,
                    'processed_data': None
                }
            
            # Step 2: Content validation and loading
            file_content = uploaded_file.read()
            uploaded_file.seek(0)  # Reset file pointer
            
            # Step 3: Load and validate Excel content
            try:
                # Use openpyxl for initial validation
                workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
                worksheet_names = workbook.sheetnames
                
                # Load data with pandas
                if len(worksheet_names) > 1:
                    # Multi-sheet workbook - let user choose or load first sheet
                    sheet_data = {}
                    for sheet_name in worksheet_names:
                        try:
                            df = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet_name, engine='openpyxl')
                            if not df.empty:
                                sheet_data[sheet_name] = df
                        except Exception as e:
                            continue
                    
                    if sheet_data:
                        # Use the first non-empty sheet
                        first_sheet = list(sheet_data.keys())[0]
                        main_df = sheet_data[first_sheet]
                        st.info(f"📊 Loaded sheet: '{first_sheet}' from {len(worksheet_names)} available sheets")
                    else:
                        return {
                            'success': False,
                            'message': 'No valid data found in any worksheet',
                            'data': None,
                            'processed_data': None
                        }
                else:
                    # Single sheet workbook
                    main_df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
                
                workbook.close()
                
            except Exception as e:
                return {
                    'success': False,
                    'message': f'Failed to read Excel file: {str(e)}',
                    'data': None,
                    'processed_data': None
                }
            
            # Step 4: Data validation and cleaning
            processed_data = self._process_data(main_df)
            
            # Step 5: Security scan of processed data
            security_check = self._security_scan(main_df)
            if not security_check['safe']:
                return {
                    'success': False,
                    'message': f'Security check failed: {security_check["message"]}',
                    'data': None,
                    'processed_data': None
                }
            
            return {
                'success': True,
                'message': 'File processed successfully',
                'data': main_df,
                'processed_data': processed_data
            }
            
        except Exception as e:
            return {
                'success': False,
                'message': f'Unexpected error: {str(e)}',
                'data': None,
                'processed_data': None
            }
    
    def _validate_file(self, uploaded_file) -> Dict[str, Any]:
        """Validate file properties and security constraints"""
        
        # Check file size
        if uploaded_file.size > self.MAX_FILE_SIZE:
            return {
                'valid': False,
                'message': f'File size ({uploaded_file.size / (1024*1024):.1f}MB) exceeds limit ({self.MAX_FILE_SIZE / (1024*1024)}MB)'
            }
        
        # Check file extension
        file_extension = '.' + uploaded_file.name.split('.')[-1].lower()
        if file_extension not in self.ALLOWED_EXTENSIONS:
            return {
                'valid': False,
                'message': f'File type "{file_extension}" not allowed. Only .xlsx files are supported.'
            }
        
        # Check MIME type
        if uploaded_file.type not in self.ALLOWED_MIME_TYPES:
            return {
                'valid': False,
                'message': f'Invalid MIME type: {uploaded_file.type}'
            }
        
        # Validate filename (prevent directory traversal)
        if not self._is_safe_filename(uploaded_file.name):
            return {
                'valid': False,
                'message': 'Invalid filename. Please use only alphanumeric characters, spaces, hyphens, and underscores.'
            }
        
        # Check for empty file
        if uploaded_file.size == 0:
            return {
                'valid': False,
                'message': 'File appears to be empty'
            }
        
        return {'valid': True, 'message': 'File validation successful'}
    
    def _is_safe_filename(self, filename: str) -> bool:
        """Check if filename is safe (prevent directory traversal and injection)"""
        # Remove path components and check for dangerous characters
        safe_pattern = re.compile(r'^[a-zA-Z0-9._\\-\\s()]+$')
        
        # Check for directory traversal attempts
        if '..' in filename or '/' in filename or '\\\\' in filename:
            return False
        
        # Check against safe pattern
        return bool(safe_pattern.match(filename))
    
    def _process_data(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Process and analyze the loaded data"""
        processed_data = {
            'columns': df.columns.tolist(),
            'data_types': df.dtypes.to_dict(),
            'shape': df.shape,
            'missing_values': df.isnull().sum().to_dict(),
            'numeric_columns': df.select_dtypes(include=['number']).columns.tolist(),
            'categorical_columns': df.select_dtypes(include=['object', 'category']).columns.tolist(),
            'datetime_columns': df.select_dtypes(include=['datetime64']).columns.tolist()
        }
        
        # Generate sample values for each column (for chatbot context)
        processed_data['sample_values'] = {}
        for col in df.columns:
            try:
                unique_vals = df[col].dropna().unique()
                if len(unique_vals) > 0:
                    processed_data['sample_values'][col] = unique_vals[:5].tolist()
                else:
                    processed_data['sample_values'][col] = []
            except Exception:
                processed_data['sample_values'][col] = []
        
        # Generate basic statistics for numeric columns
        if processed_data['numeric_columns']:
            try:
                stats = df[processed_data['numeric_columns']].describe().to_dict()
                processed_data['statistics'] = stats
            except Exception:
                processed_data['statistics'] = {}
        
        return processed_data
    
    def _security_scan(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Perform security scanning on the data content"""
        
        # Check for suspicious patterns in string columns
        suspicious_patterns = [
            r'<script[^>]*>.*?</script>',  # XSS attempts
            r'javascript:',                 # JavaScript injection
            r'vbscript:',                  # VBScript injection
            r'on\\w+\\s*=',                 # Event handlers
            r'SELECT.*FROM',               # SQL injection (basic)
            r'DROP\\s+TABLE',               # SQL injection (destructive)
            r'INSERT\\s+INTO',              # SQL injection
            r'UPDATE\\s+.*SET',             # SQL injection
            r'DELETE\\s+FROM',              # SQL injection
        ]
        
        try:
            string_columns = df.select_dtypes(include=['object']).columns
            
            for col in string_columns:
                col_data = df[col].dropna().astype(str)
                
                for pattern in suspicious_patterns:
                    if col_data.str.contains(pattern, case=False, regex=True, na=False).any():
                        return {
                            'safe': False,
                            'message': f'Suspicious content detected in column "{col}". Please review your data.'
                        }
            
            # Check for excessively large cells (potential DoS)
            max_cell_size = 10000  # 10KB per cell
            for col in df.columns:
                if df[col].dtype == 'object':
                    cell_sizes = df[col].dropna().astype(str).str.len()
                    if cell_sizes.max() > max_cell_size:
                        return {
                            'safe': False,
                            'message': f'Cell content in column "{col}" exceeds size limit'
                        }
            
            return {'safe': True, 'message': 'Security scan passed'}
            
        except Exception as e:
            # If security scan fails, err on the side of caution
            return {
                'safe': False,
                'message': f'Security scan failed: {str(e)}'
            }
    
    @staticmethod
    def generate_file_hash(file_content: bytes) -> str:
        """Generate SHA-256 hash of file content for integrity checking"""
        return hashlib.sha256(file_content).hexdigest()
```
